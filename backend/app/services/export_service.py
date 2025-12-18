from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.models import ExportJob, ProductRecord
from app.services.audit_service import AuditService
from app.services.log_service import LogService
from app.utils.time import utc_now
from app.models.extraction_run import ExtractionRun

ALLOWED_EXPORT_FIELDS = {"asin", "title", "price", "batch_id", "validation_status", "currency"}


@dataclass
class ExportResult:
    file_path: Path
    rows: int


class ExportService:
    def __init__(self) -> None:
        self.export_dir = settings.storage_dir / "exports"
        self.export_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def normalize_export_type(export_type: str) -> str:
        mapping = {
            "clean-products": "clean_products",
            "clean_products": "clean_products",
            "failed-rows": "failed_rows",
            "failed_rows": "failed_rows",
            "extraction_results": "extraction_results",
        }
        normalized = mapping.get(export_type)
        if not normalized:
            raise ValueError("不支持的导出类型")
        return normalized

    @staticmethod
    def normalize_file_format(file_format: str) -> str:
        fmt = (file_format or "csv").lower()
        if fmt not in {"csv", "xlsx"}:
            raise ValueError("不支持的导出文件格式")
        return fmt

    @staticmethod
    def _validate_fields(selected_fields: list[str], export_type: str = "") -> list[str]:
        if export_type == "extraction_results":
            # For extraction results, we allow any fields (dynamic)
            return selected_fields
            
        fields = [f for f in selected_fields if f in ALLOWED_EXPORT_FIELDS]
        if not fields:
            raise ValueError("请选择至少一个允许导出的字段")
        return fields

    def create_job(
        self,
        db: Session,
        *,
        export_type: str,
        filters: dict | None,
        selected_fields: list[str],
        file_format: str,
        triggered_by: str | None = None,
    ) -> ExportJob:
        normalized_type = self.normalize_export_type(export_type)
        normalized_format = self.normalize_file_format(file_format)
        job = ExportJob(
            export_type=normalized_type,
            filters=filters,
            selected_fields=self._validate_fields(selected_fields, normalized_type),
            file_format=normalized_format,
            status="running",
            started_at=utc_now(),
            triggered_by=triggered_by,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        try:
            result = self._generate_file(db, job)
            job.file_path = str(result.file_path)
            job.status = "succeeded"
            job.finished_at = utc_now()
            job.error_message = None
            LogService.log(
                db,
                level="info",
                category="export",
                message=f"导出完成，任务 {job.id}",
                context={
                    "export_type": job.export_type,
                    "file_format": job.file_format,
                    "rows": result.rows,
                    "file_path": job.file_path,
                },
                trace_id=job.id,
            )
        except Exception as exc:  # noqa: BLE001
            job.status = "failed"
            job.error_message = str(exc)
            job.finished_at = utc_now()
            LogService.log(
                db,
                level="error",
                category="export",
                message=f"导出失败，任务 {job.id}",
                context={
                    "export_type": job.export_type,
                    "file_format": job.file_format,
                    "error": str(exc),
                },
                trace_id=job.id,
            )
        finally:
            db.add(job)
            db.commit()
            db.refresh(job)
            AuditService.log_action(
                db,
                action="export.create",
                actor_id=triggered_by,
                entity_id=job.id,
                payload={"status": job.status, "file_path": job.file_path},
            )
        return job

    def _generate_file(self, db: Session, job: ExportJob) -> ExportResult:
        rows = self._fetch_rows(db, job.export_type, job.filters)
        filename = self.export_dir / f"{job.id}.{job.file_format}"
        
        # Determine fields to write
        fields = job.selected_fields
        
        # Special handling for extraction_results:
        # User wants "Source Data + Selected AI Fields".
        # So we should ALWAYS include source fields (Standard + Raw), 
        # and only filter AI fields based on selection.
        if job.export_type == "extraction_results" and rows:
            # 1. Get AI fields definition from the run
            run_id = job.filters.get("run_id")
            ai_fields = set()
            if run_id:
                run = db.get(ExtractionRun, run_id)
                if run and run.target_fields:
                    ai_fields = set(run.target_fields)
            
            # 2. Identify all available keys from data
            all_keys = list(rows[0].keys())
            
            # 3. Reorder fields: Essentials -> AI Fields -> Others
            essential_keys = ['asin', 'title', 'bullets', 'feature_bullets', 'price', 'brand']
            
            final_fields = []
            processed_keys = set()
            
            # 3.1 Essentials
            for key in essential_keys:
                if key in all_keys:
                    final_fields.append(key)
                    processed_keys.add(key)
            
            # 3.2 AI Fields (Use order from run.target_fields if available)
            ai_ordered = []
            if run and run.target_fields:
                ai_ordered = run.target_fields
            else:
                ai_ordered = sorted(list(ai_fields))
                
            for key in ai_ordered:
                if key in all_keys and key not in processed_keys:
                    # Check selection
                    if not job.selected_fields or key in job.selected_fields:
                        final_fields.append(key)
                        processed_keys.add(key)
            
            # 3.3 Others (Raw fields not in essentials, and any AI fields missed)
            for key in all_keys:
                if key not in processed_keys:
                    if key in ai_fields:
                        if not job.selected_fields or key in job.selected_fields:
                            final_fields.append(key)
                    else:
                        final_fields.append(key)
            
            fields = final_fields

        if not fields and rows:
            # If no fields selected (or dynamic), use keys from first row
            fields = list(rows[0].keys())
            
        if job.file_format == "csv":
            with filename.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=fields)
                writer.writeheader()
                for item in rows:
                    writer.writerow({field: item.get(field) for field in fields})
        else:
            # Real XLSX generation with styling
            ai_columns = set()
            if job.export_type == "extraction_results":
                run_id = job.filters.get("run_id")
                if run_id:
                    run = db.get(ExtractionRun, run_id)
                    if run and run.target_fields:
                        ai_columns = set(run.target_fields)
            
            self._generate_xlsx(filename, fields, rows, job, ai_columns)
        return ExportResult(file_path=filename, rows=len(rows))
    
    def _generate_xlsx(self, filename: Path, fields: list[str], rows: list[dict], job: ExportJob, ai_columns: set[str] = None) -> None:
        """Generate XLSX file with styling for AI columns."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Data"
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        # Styles
        # Header Styles
        header_font = Font(bold=True, size=11, color="000000")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Standard Column Header (Light Blue)
        standard_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
        # AI Column Header (Light Orange)
        ai_fill = PatternFill(start_color="FFF7E6", end_color="FFF7E6", fill_type="solid")
        
        # Cell Styles
        # Disable wrap_text for data cells to keep row height compact
        cell_align = Alignment(vertical="center", wrap_text=False)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Write header row
        ws.row_dimensions[1].height = 30  # Taller header
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            
            if ai_columns and field in ai_columns:
                cell.fill = ai_fill
            else:
                cell.fill = standard_fill
        
        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            # Set fixed row height for data rows (optional, but good for consistency)
            ws.row_dimensions[row_idx].height = 20
            for col_idx, field in enumerate(fields, start=1):
                value = row_data.get(field)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx, field in enumerate(fields, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            # Set specific widths for known columns
            if field.lower() == 'asin':
                ws.column_dimensions[column_letter].width = 15
            elif field.lower() in ('price', 'rating', 'reviews', 'currency'):
                ws.column_dimensions[column_letter].width = 12
            elif field.lower() in ('title', 'title_cn'):
                ws.column_dimensions[column_letter].width = 50
            elif field.lower() in ('bullets', 'bullets_cn'):
                ws.column_dimensions[column_letter].width = 60
            elif ai_columns and field in ai_columns:
                # AI columns usually contain text, give them space
                ws.column_dimensions[column_letter].width = 40
            else:
                # Auto-width for others (capped)
                max_length = len(str(field)) + 4
                for row_data in rows[:50]:  # Sample first 50 rows
                    cell_value = str(row_data.get(field, ""))
                    # Cap sample length to avoid huge calculation for long text
                    length = min(len(cell_value), 50)
                    max_length = max(max_length, length)
                
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        wb.save(filename)

    def _fetch_rows(self, db: Session, export_type: str, filters: dict | None) -> Sequence[dict]:
        if export_type == "extraction_results":
            return self._fetch_extraction_rows(db, filters)

        stmt = select(ProductRecord)
        if filters:
            batch_id = filters.get("batch_id")
            if batch_id:
                # Convert to int if it's a string
                batch_id_int = int(batch_id) if isinstance(batch_id, str) else batch_id
                stmt = stmt.where(ProductRecord.batch_id == batch_id_int)
            status = filters.get("validation_status")
            if status:
                stmt = stmt.where(ProductRecord.validation_status == status)
        records = db.execute(stmt).scalars().all()
        return [
            {
                "batch_id": item.batch_id,
                "asin": item.asin,
                "title": item.title,
                "price": item.price,
                "validation_status": item.validation_status,
            }
            for item in records
        ]

    def _fetch_extraction_rows(self, db: Session, filters: dict | None) -> Sequence[dict]:
        if not filters or not filters.get("run_id"):
            raise ValueError("导出提取结果需要指定 run_id")
            
        run_id = filters.get("run_id")
        run = db.get(ExtractionRun, run_id)
        if not run:
            raise ValueError("找不到指定的提取记录")
            
        stmt = select(ProductRecord).where(ProductRecord.batch_id == run.batch_id)
        records = db.execute(stmt).scalars().all()
        
        results = []
        for record in records:
            # 1. Start with Raw Payload (Original Data)
            # This ensures we preserve ALL original columns exactly as uploaded
            row = record.raw_payload.copy() if record.raw_payload else {}
            
            # 2. Merge AI Features
            if record.ai_features:
                for k, v in record.ai_features.items():
                    # Skip metadata fields
                    if k.startswith("_"):
                        continue
                    row[k] = v
            
            # 3. Sanitize all values for Excel (convert dict/list to JSON string)
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    import json
                    row[k] = json.dumps(v, ensure_ascii=False)
                
            results.append(row)
            
        return results

    def list_jobs(self, db: Session, limit: int = 20, offset: int = 0) -> list[ExportJob]:
        """获取导出任务列表,按开始时间倒序"""
        stmt = select(ExportJob).order_by(ExportJob.started_at.desc()).limit(limit).offset(offset)
        return list(db.execute(stmt).scalars().all())

    def get_job(self, db: Session, job_id: str) -> ExportJob | None:
        return db.get(ExportJob, job_id)


export_service = ExportService()
