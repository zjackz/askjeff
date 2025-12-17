from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import settings
from app.models.import_batch import ProductRecord
from app.services.audit_service import AuditService
from app.services.import_config import ImportConfig, load_import_config
from app.services.import_repository import ImportRepository
from app.services.log_service import LogService

STRATEGY_ALIASES = {
    "overwrite": "overwrite",
    "append": "append",
    "update-only": "update_only",
    "update_only": "update_only",
}

STANDARD_FIELDS = {
    "asin",
    "title",
    "category",
    "price",
    "currency",
    "sales_rank",
    "reviews",
    "rating",
}


@dataclass
class ParsedResult:
    records: list[ProductRecord]
    failures: list[dict]
    warnings_count: int
    columns_seen: list[str]
    columns_mapped: list[str]
    columns_unmapped: list[str]


class ImportAbort(Exception):
    """用于中断导入的异常。"""


class ImportService:
    def __init__(self) -> None:
        self.import_dir = settings.storage_dir / "imports"
        self.failed_dir = settings.storage_dir / "exports" / "failed"
        self.import_dir.mkdir(parents=True, exist_ok=True)
        self.failed_dir.mkdir(parents=True, exist_ok=True)
        self.base_config: ImportConfig = load_import_config()

    @staticmethod
    def normalize_strategy(strategy: str) -> str:
        """将契约中的导入策略转换为内部枚举值。"""
        key = strategy.strip().lower()
        normalized = STRATEGY_ALIASES.get(key)
        if not normalized:
            raise ValueError("不支持的导入策略")
        return normalized

    def handle_upload(
        self,
        db: Session,
        *,
        file: UploadFile,
        import_strategy: str,
        sheet_name: str | None = None,
        on_missing_required: str | None = None,
        column_aliases: dict[str, str] | None = None,
        created_by: str | None = None,
    ):
        saved_path = self._save_file(file)
        
        # Calculate hash
        file_hash = self._calculate_file_hash(saved_path)
        
        # Check for duplicates
        # existing_batch = ImportRepository.find_batch_by_hash(db, file_hash)
        # if existing_batch:
        #     # Clean up uploaded file
        #     saved_path.unlink(missing_ok=True)
        #     raise ValueError(f"检测到重复文件导入 (与批次 {existing_batch.id} 内容相同)")

        effective_config = self.base_config.merge_overrides(
            sheet_name=sheet_name,
            on_missing_required=on_missing_required,
            column_aliases=column_aliases,
        )
        batch = ImportRepository.create_batch(
            db,
            filename=file.filename,
            storage_path=str(saved_path),
            import_strategy=import_strategy,
            sheet_name=effective_config.sheet_name,
            created_by=created_by,
            file_hash=file_hash,
        )
        
        # Update status to running
        ImportRepository.update_batch_progress(db, batch, status="running", total_rows=0)

        failure_file = None
        try:
            parsed = self._parse_file(saved_path, batch_id=batch.id, config=effective_config)
            if parsed.records:
                ImportRepository.create_product_records(db, parsed.records)
            if parsed.failures:
                failure_file = self._write_failures(batch.id, parsed.failures)
            status = "succeeded" if not parsed.failures else "failed"
            failure_summary = {
                "columns_seen": parsed.columns_seen,
                "columns_mapped": parsed.columns_mapped,
                "columns_unmapped": parsed.columns_unmapped,
                "warnings_count": parsed.warnings_count,
                "failed_rows_path": str(failure_file.relative_to(settings.storage_dir))
                if failure_file
                else None,
                "items": parsed.failures if parsed.failures else None,
            }
            ImportRepository.update_batch_stats(
                db,
                batch,
                status=status,
                total_rows=len(parsed.records) + len(parsed.failures),
                success_rows=len(parsed.records),
                failed_rows=len(parsed.failures),
                failure_summary=failure_summary,
                columns_seen=parsed.columns_seen,
            )
            LogService.log(
                db,
                level="info",
                category="import",
                message=f"导入完成，批次 {batch.id}，状态 {status}",
                context={
                    "filename": file.filename,
                    "status": status,
                    "total": len(parsed.records) + len(parsed.failures),
                    "success": len(parsed.records),
                    "failed": len(parsed.failures),
                    "sheet": effective_config.sheet_name,
                },
                trace_id=batch.id,
            )
        except ImportAbort as exc:
            ImportRepository.update_batch_stats(
                db,
                batch,
                status="failed",
                total_rows=0,
                success_rows=0,
                failed_rows=0,
                failure_summary={"error": str(exc)},
                columns_seen=[],
            )
            LogService.log(
                db,
                level="error",
                category="import",
                message=f"导入失败，批次 {batch.id}",
                context={"error": str(exc), "filename": file.filename},
                trace_id=batch.id,
            )
            raise ValueError(str(exc))

        AuditService.log_action(
            db,
            action="import.create",
            actor_id=created_by,
            entity_id=batch.id,
            payload={
                "filename": file.filename,
                "status": batch.status,
                "failed_rows": batch.failed_rows,
                "sheet_name": effective_config.sheet_name,
            },
        )
        return batch

    def _save_file(self, file: UploadFile) -> Path:
        suffix = Path(file.filename or "upload").suffix or ".csv"
        target = self.import_dir / f"{uuid4()}{suffix}"
        with target.open("wb") as dest:
            dest.write(file.file.read())
        file.file.seek(0)
        return target

    def _calculate_file_hash(self, path: Path) -> str:
        sha256_hash = hashlib.sha256()
        with path.open("rb") as f:
            # Read and update hash string value in blocks of 4K
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

    def _parse_file(self, path: Path, *, batch_id: str, config: ImportConfig) -> ParsedResult:
        records: list[ProductRecord] = []
        failures: list[dict] = []
        warnings_count = 0

        rows = self._read_rows(path, sheet_name=config.sheet_name)
        if not rows:
            raise ImportAbort("指定 sheet 为空或不存在数据")

        # 部分 Excel 首行会放置备注/提示，此时第二行才是真正表头
        header_idx = 0
        if len(rows) > 1:
            non_empty = sum(1 for cell in rows[0] if cell not in (None, ""))
            if non_empty < 2:
                header_idx = 1
        headers = rows[header_idx]
        data_rows = rows[header_idx + 1 :]

        columns_seen = [str(h).strip() for h in headers if h is not None]
        header_keyed = [h.strip() if isinstance(h, str) else str(h) for h in headers]
        alias_map = {k.lower(): v for k, v in config.column_aliases.items()}

        columns_mapped = set()
        columns_unmapped = set()

        for idx, raw_values in enumerate(data_rows, start=header_idx + 2):
            row_dict: dict[str, Any] = {
                header_keyed[i]: self._to_jsonable(raw_values[i]) 
                for i in range(len(header_keyed))
            }
            if all(value in (None, "") for value in row_dict.values()):
                continue
            mapped_payload: dict[str, Any] = {}
            normalized_payload: dict[str, Any] = {}
            validation_messages: dict[str, str] = {}
            validation_status = "valid"

            for header, value in row_dict.items():
                key_lower = header.lower()
                std_field = alias_map.get(key_lower)
                if std_field is None and key_lower in STANDARD_FIELDS:
                    std_field = key_lower
                if std_field:
                    columns_mapped.add(std_field)
                    mapped_payload[std_field] = value
                    normalized_value, warn = self._normalize_value(std_field, value, config)
                    if warn:
                        validation_status = "warning"
                        validation_messages[std_field] = warn
                        warnings_count += 1
                    normalized_payload[std_field] = self._to_jsonable(normalized_value)
                else:
                    columns_unmapped.add(header)

            default_curr = config.normalization.get("default_currency")
            if "currency" not in mapped_payload and default_curr:
                mapped_payload["currency"] = default_curr
                normalized_payload["currency"] = default_curr
                columns_mapped.add("currency")

            missing = [field for field in config.required_fields if not mapped_payload.get(field)]
            if missing:
                reason = f"缺少必填字段: {', '.join(missing)}"
                failures.append(
                    {
                        "rowNumber": idx,
                        "asin": mapped_payload.get("asin"),
                        "reason": reason,
                        "rawValues": row_dict,
                    }
                )
                if config.on_missing_required == "abort":
                    raise ImportAbort(reason)
                continue

            # 使用统一标准化器
            from app.services.product_normalizer import ProductDataNormalizer
            from app.core.logger import logger
            
            try:
                # 1. 标准化数据
                normalized = ProductDataNormalizer.normalize_product(
                    raw_data=row_dict,
                    source="file"
                )
                
                # 2. 验证数据
                validation_status, validation_messages = ProductDataNormalizer.validate_product(
                    normalized
                )
                
                # 统计警告
                if validation_status == "warning":
                    warnings_count += 1
                elif validation_status == "error":
                    if config.on_missing_required == "abort":
                        raise ImportAbort(f"数据验证失败: {validation_messages}")
                    failures.append({
                        "rowNumber": idx,
                        "asin": normalized.get("asin"),
                        "reason": f"验证失败: {validation_messages}",
                        "rawValues": row_dict,
                    })
                    continue
                
                # 3. 创建 normalized_payload
                normalized_payload = ProductDataNormalizer.create_normalized_payload(normalized)
                
                # 4. 创建记录
                record = ProductRecord(
                    batch_id=batch_id,
                    asin=normalized["asin"],
                    title=normalized["title"],
                    category=normalized["category"],
                    price=normalized["price"],
                    currency=normalized["currency"],
                    sales_rank=normalized["sales_rank"],
                    reviews=normalized["reviews"],
                    rating=normalized["rating"],
                    raw_payload=normalized["raw_payload"],
                    normalized_payload=normalized_payload,
                    extended_data=normalized.get("extended_data"),  # 新增
                    data_source=normalized.get("data_source", "file"),  # 新增
                    validation_status=validation_status,
                    validation_messages=validation_messages,
                )
                records.append(record)
                
            except Exception as e:
                logger.error(f"处理行 {idx} 失败: {e}", exc_info=True)
                failures.append({
                    "rowNumber": idx,
                    "asin": row_dict.get("asin") or row_dict.get("ASIN"),
                    "reason": f"处理失败: {str(e)}",
                    "rawValues": row_dict,
                })
                continue

        return ParsedResult(
            records=records,
            failures=failures,
            warnings_count=warnings_count,
            columns_seen=list(columns_seen),
            columns_mapped=sorted(columns_mapped),
            columns_unmapped=sorted(columns_unmapped),
        )

    def _read_rows(self, path: Path, *, sheet_name: str) -> list[list[Any]]:
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                # Fallback: if only one sheet, use it
                if len(wb.sheetnames) == 1:
                    ws = wb[wb.sheetnames[0]]
                else:
                    raise ImportAbort(f"未找到指定 sheet: {sheet_name}，可用 sheet: {', '.join(wb.sheetnames)}")
            else:
                ws = wb[sheet_name]
            return [list(row) for row in ws.iter_rows(values_only=True)]
        
        # CSV encoding detection
        encodings = ["utf-8", "gb18030"]  # gb18030 covers gbk and gb2312
        for enc in encodings:
            try:
                with path.open("r", encoding=enc, newline="") as fh:
                    reader = csv.reader(fh)
                    rows = [row for row in reader]
                return rows
            except UnicodeDecodeError:
                continue
            except Exception as e:
                raise ImportAbort(f"读取 CSV 失败: {e}")
        
        raise ImportAbort("无法识别文件编码，请使用 UTF-8 或 GBK/GB18030 编码")

    def _write_failures(self, batch_id: str, failures: list[dict]) -> Path:
        target = self.failed_dir / f"{batch_id}_failed.csv"
        with target.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.DictWriter(
                fh, fieldnames=["rowNumber", "asin", "reason", "rawValues"], extrasaction="ignore"
            )
            writer.writeheader()
            for item in failures:
                writer.writerow(
                    {
                        "rowNumber": item.get("rowNumber"),
                        "asin": item.get("asin"),
                        "reason": item.get("reason"),
                        "rawValues": item.get("rawValues"),
                    }
                )
        return target

    def _normalize_value(self, field: str, value: Any, config: ImportConfig) -> tuple[Any, str | None]:
        """按字段归一化，返回 (值, 警告信息)。"""
        if value is None or value == "":
            return None, None
        if field in {"price"}:
            dec = self._to_decimal(value, scale=int(config.normalization.get("price_scale", 2)))
            if dec is None:
                return None, "价格解析失败"
            return dec, None
        if field in {"rating"}:
            dec = self._to_decimal(value, scale=int(config.normalization.get("rating_scale", 2)))
            if dec is None:
                return None, "评分解析失败"
            return dec, None
        if field in {"sales_rank", "reviews"}:
            iv = self._to_int(value)
            if iv is None:
                return None, f"{field} 解析失败"
            return iv, None
        return value, None

    @staticmethod
    def _to_decimal(value: Any, scale: int | None = None) -> Decimal | None:
        if value in (None, ""):
            return None
        try:
            dec = Decimal(str(value))
            if scale is not None:
                return dec.quantize(Decimal(10) ** -scale)
            return dec
        except (InvalidOperation, ValueError):
            return None

    @staticmethod
    def _to_int(value: Any) -> int | None:
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _normalize_currency(value: Any, config: ImportConfig) -> str | None:
        if value in (None, ""):
            return None
        curr = str(value).upper()
        whitelist = config.normalization.get("currency_whitelist") or []
        if whitelist and curr not in whitelist:
            return None
        return curr

    @staticmethod
    def _to_jsonable(value: Any) -> Any:
        """将 Decimal 等不可 JSON 序列化的对象转换为基础类型。"""
        from datetime import date, datetime
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return value


import_service = ImportService()
